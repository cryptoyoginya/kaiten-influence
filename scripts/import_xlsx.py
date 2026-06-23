#!/usr/bin/env python3
"""Импорт исходного Excel-лонглиста в чистый нормализованный датасет.

Источник: 13 листов (ниши + черновики + ШОРТЛИСТ + Неделя 1).
Выход:
  data/channels.json   — все блогеры, дедуп по telegram-ссылке, ниша из листа
  data/sprints.json    — спринты (пока один: Неделя 1) + размещения с пайплайном

Запуск: python3 scripts/import_xlsx.py "<путь к .xlsx>"
"""
import json, re, sys, unicodedata
from pathlib import Path
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else \
    str(Path.home() / "Downloads" / "❗️1 волна_лонглист_-2.xlsx")
OUT = Path(__file__).resolve().parent.parent / "data"
OUT.mkdir(exist_ok=True)

# Лист -> (ниша, draft?). Имена с хвостовыми пробелами как в книге.
NICHE_SHEETS = {
    "Бизнес ": ("Бизнес", False),
    "Продакты и CEO": ("Продакты и CEO", False),
    "PM и управление": ("PM и управление", False),
    "CTO и тимлид ": ("CTO и тимлид", False),
    "Консалтинг ": ("Консалтинг", False),
    "Производство ": ("Производство", False),
    "Финансы черновик": ("Финансы", True),
    "Агентства и медиа черновик ": ("Агентства и медиа", True),
    "Бизнес (вне ниш) черновик": ("Бизнес (вне ниш)", True),
    "Производство черновик": ("Производство", True),
    "PM и управление черновик ": ("PM и управление", True),
}
SPRINT_SHEET = "1️⃣ Неделя 1 22-26 июня "
SHORTLIST_SHEET = "ШОРТЛИСТ"


def clean(v):
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    return s


# Заголовки-разделители из черновиков — не каналы.
JUNK_NAMES = {"крупные каналы", "средние каналы", "малые каналы",
              "итого", "всего"}


def handle_to_link(name, link):
    """Если ссылки нет, но в названии есть @handle — собираем t.me-ссылку."""
    if clean(link):
        return clean(link)
    m = re.search(r"@([a-zA-Z0-9_]{4,})", clean(name))
    if m:
        return "https://t.me/" + m.group(1)
    return ""


def slug_link(link):
    """Ключ дедупа: нормализованный t.me-хэндл, иначе пусто."""
    s = clean(link).lower()
    m = re.search(r"t\.me/([a-z0-9_]+)", s)
    if m:
        return "tg:" + m.group(1)
    m = re.search(r"youtube\.com/([^\s/?]+)", s)
    if m:
        return "yt:" + m.group(1)
    return ""


def name_key(name):
    s = clean(name).lower()
    s = unicodedata.normalize("NFKD", s)
    s = re.sub(r"[^\w]+", "", s)
    return "nm:" + s if s else ""


def header_index(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        cells = [clean(c).lower() for c in row]
        if any(c.startswith("назван") or c.startswith("блогер") for c in cells):
            return i, [clean(c) for c in row]
    return 0, [clean(c) for c in next(ws.iter_rows(max_row=1, values_only=True))]


def pick(header, row, *needles):
    for idx, h in enumerate(header):
        hl = h.lower()
        if any(n in hl for n in needles):
            if idx < len(row):
                return clean(row[idx])
    return ""


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    channels = {}   # key -> channel dict
    order = []

    def upsert(rec):
        key = slug_link(rec["link"]) or name_key(rec["name"])
        if not key:
            return
        if key in channels:
            c = channels[key]
            # дополняем пустые поля и копим ниши/комментарии
            for f in ("audience", "themes", "subscribers", "err_views", "price_raw"):
                if not c.get(f) and rec.get(f):
                    c[f] = rec[f]
            if rec["niche"] not in c["niches"]:
                c["niches"].append(rec["niche"])
            for cm in rec["comments"]:
                if cm and cm not in c["comments"]:
                    c["comments"].append(cm)
            if rec["referral"] and not c["referral"]:
                c["referral"] = rec["referral"]
            c["draft"] = c["draft"] and rec["draft"]  # не-черновик побеждает
        else:
            channels[key] = rec
            rec["niches"] = [rec.pop("niche")]
            order.append(key)

    # --- ниши ---
    for sheet, (niche, draft) in NICHE_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hi, header = header_index(ws)
        for row in ws.iter_rows(min_row=hi + 2, values_only=True):
            name = clean(row[0]) if row else ""
            if not name or name.lower() in JUNK_NAMES:
                continue
            comments = [pick(header, row, "комментарий", "коммент")]
            extra = [clean(c) for c in row[7:] if clean(c)]
            comments += [e for e in extra if e not in comments]
            upsert({
                "name": name,
                "link": handle_to_link(name, pick(header, row, "ссылк")),
                "niche": niche,
                "subscribers": pick(header, row, "подписч"),
                "audience": pick(header, row, "аудитор"),
                "themes": pick(header, row, "темы", "тематик"),
                "err_views": pick(header, row, "просмотр", "err"),
                "price_raw": pick(header, row, "стоимост", "цена"),
                "referral": pick(header, row, "рефералк"),
                "comments": [c for c in comments if c],
                "draft": draft,
                # рабочие поля размещения (заполняются в спринте), пустые по умолчанию
                "post_date": "", "post_topic": "", "offer": "",
                "creative": "", "landing": "", "utm": "",
                "shortlisted": False,
            })

    # --- шортлист: помечаем флаг ---
    if SHORTLIST_SHEET in wb.sheetnames:
        ws = wb[SHORTLIST_SHEET]
        hi, _ = header_index(ws)
        for row in ws.iter_rows(min_row=hi + 2, values_only=True):
            name = clean(row[0]) if row else ""
            link = clean(row[1]) if len(row) > 1 else ""
            if not name:
                continue
            key = slug_link(link) or name_key(name)
            if key in channels:
                channels[key]["shortlisted"] = True

    # --- спринт Неделя 1 ---
    placements = []
    if SPRINT_SHEET in wb.sheetnames:
        ws = wb[SPRINT_SHEET]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        STEPS = ["Внутреннее согласование", "Согласование с инфлом", "Реквизиты для договора",
                 "Договор готов", "Договор подписан", "Оплата",
                 "Маркировка получена", "Маркировка в посте", "Опубликовано", "Аналитика"]
        for row in rows:
            name = clean(row[0]) if row else ""
            if not name or name.upper() == "ИТОГО":
                continue
            steps = {}
            for j, label in enumerate(STEPS):
                col = 16 + j  # колонки Q.. (0-indexed 16) — чекбоксы пайплайна
                val = clean(row[col]) if col < len(row) else ""
                steps[label] = bool(val) and val.lower() not in ("нет", "0", "false")
            placements.append({
                "name": name,
                "author_desc": clean(row[1]),
                "audience": clean(row[2]),
                "post_date": clean(row[3]),
                "post_topic": clean(row[4]),
                "offer": clean(row[5]),
                "creative": clean(row[6]),
                "landing": clean(row[7]),
                "utm": clean(row[8]),
                "price": clean(row[9]),
                "price_discount": clean(row[10]),
                "subscribers": clean(row[11]),
                "avg_views": clean(row[12]),
                "err": clean(row[13]),
                "forecast_reach": clean(row[14]),
                "forecast_cpv": clean(row[15]),
                "steps": steps,
            })

    sprints = [{
        "id": "week-1",
        "title": "Неделя 1",
        "date_from": "2026-06-22",
        "date_to": "2026-06-26",
        "status": "active",
        "placements": placements,
    }]

    chan_list = [channels[k] for k in order]

    # --- интеграции: засев из размещений спринта, поля результата пустые ---
    # ниша берётся из канала по имени, если нашёлся
    niche_by_name = {}
    for c in chan_list:
        niche_by_name[name_key(c["name"])] = c["niches"][0] if c["niches"] else ""

    def empty_result():
        return {
            "post_link": "",
            "format": "",
            "costs": {"price": "", "marking": "", "tax": "", "total": ""},
            "reach": {"views": "", "reach": "", "likes": "", "reposts": "",
                       "comments_count": "", "er": ""},
            "conversion": {"clicks": "", "registrations": "", "activations": "",
                            "paying": "", "revenue": ""},
            "unit": {"cpv": "", "cpm": "", "ctr": "", "cpl": "", "cac": "",
                      "romi": "", "payback": ""},
            "screens": {"creative": "", "stats": "", "comments": []},
            "lessons": {"sentiment": "", "worked": "", "failed": "",
                         "learned": "", "verdict": ""},
        }

    integrations = []
    for sp in sprints:
        for p in sp["placements"]:
            integrations.append({
                "id": f"{sp['id']}-{name_key(p['name'])}",
                "sprint_id": sp["id"],
                "name": p["name"],
                "niche": niche_by_name.get(name_key(p["name"]), ""),
                "date": p["post_date"],
                "landing": p["landing"],
                "published": bool(p["steps"].get("Опубликовано")),
                # бриф размещения — редактируемые данные карточки
                "brief": {
                    "author_desc": p["author_desc"],
                    "audience": p["audience"],
                    "date": p["post_date"],
                    "post_topic": p["post_topic"],
                    "offer": p["offer"],
                    "creative": p["creative"],
                    "landing": p["landing"],
                    "utm": p["utm"],
                },
                # план (из спринта) — для сравнения план→факт
                "plan": {
                    "price": p["price_discount"] or p["price"],
                    "reach": p["forecast_reach"],
                    "cpv": p["forecast_cpv"],
                    "err": p["err"],
                    "views": p["avg_views"],
                },
                "result": empty_result(),
            })

    (OUT / "channels.json").write_text(
        json.dumps(chan_list, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "sprints.json").write_text(
        json.dumps(sprints, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "integrations.json").write_text(
        json.dumps(integrations, ensure_ascii=False, indent=2), encoding="utf-8")

    niche_counts = {}
    for c in chan_list:
        for n in c["niches"]:
            niche_counts[n] = niche_counts.get(n, 0) + 1
    print(f"channels: {len(chan_list)} (dedup)")
    print(f"shortlisted: {sum(1 for c in chan_list if c['shortlisted'])}")
    print(f"sprint placements: {len(placements)}")
    print(f"integrations (засев): {len(integrations)}")
    print("по нишам:")
    for n, k in sorted(niche_counts.items(), key=lambda x: -x[1]):
        print(f"  {k:3}  {n}")


if __name__ == "__main__":
    main()
